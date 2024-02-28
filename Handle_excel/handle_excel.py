# -*- coding: utf-8 -*-
import time

import pandas as pd
import openpyxl

from openpyxl import Workbook, load_workbook

old = load_workbook("总数据.xlsx")
new = load_workbook("新数据.xlsx")
# print(wb)
sheet_old = old.get_sheet_by_name('Sheet1')
sheet_new = new.get_sheet_by_name('Sheet1')
list_old = []
list_new = []
for old_cell in sheet_old:
    list_old.append(old_cell[0].value)
for new_cell in sheet_new:
    list_new.append(new_cell[0].value)

new_list = list(set(list_new) - set(list_old))
print("新增数据：{}".format(new_list))
# print(list)

# 获取最大行数
max_row = sheet_old.max_row
# 新行的行号为最大行数+1
new_row = max_row + 1
for row, value in enumerate(new_list, start=1):
    sheet_old.cell(row=new_row, column=1, value=value)
    new_row += 1
    # print(new_row, value)
    time.sleep(0.5)

# print(sheet_old)
for old_cell in sheet_old:
    # print(old_cell[0].value)
    for new_cell in sheet_new:
        if old_cell[0].value == new_cell[0].value:
            # print(new_cell[0].value)
            # print(old_cell[1].value)
            # print(new_cell[1].value)
            old_cell[1].value = new_cell[1].value


old.save("替换后的数据.xlsx")
